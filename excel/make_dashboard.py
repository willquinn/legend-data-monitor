"""
Excel dashboard generator.
--------------------------------------------------------------------------------

Call `make_excel(strings, periods, data, output_path)`

================================================================================
INPUTS
================================================================================

1. strings
----------
Type:   dict[int, list[tuple[str, float]]]
Format: { string_number: [(ged_name, mass_g), ...], ... }
detectors are listed in the order of position

Example:
    strings = {
        1: [("V14654A", 3383), ("V14673A", 3450)],
        4: [("B00032C", 743),  ("V07302A", 1803)],
    }

--------------------------------------------------------------------------------

2. periods
----------
Which (run_type, run) columns exist for each period, in display order.
Controls how many runs appear in the sheet — just include what you want.

Type:   dict[str, list[tuple[str, str]]]
Format: { period: [(run_type, run), ...], ... }

Note:
  - run_type  : "cal" or "phy"
  - The last run of each period is (should be) cal-only (no trailing phy entry)
  - Remember phy runs require a cal–phy–cal sandwich to be valid
  - can/should be linked with run lists

Example (two periods, p16 with 2 runs, p18 with 1 run):
    periods = {
        "p16": [
            ("cal", "r000"), ("phy", "r000"),
            ("cal", "r001"),                   # last run — cal only
        ],
        "p18": [
            ("cal", "r000"),                   # last run — cal only
        ],
    }

--------------------------------------------------------------------------------

3. data
-------
The usability value for every (detector × column) cell.

Type:   dict[tuple, any]
Format: { (string_num, ged_name, period, run, run_type, usability_type): value }

Note:
  - usability_type  : str — "E" (energy scale) or "P" (PSD)
  - value           : what you want to display (number, string, None)
                      None leaves the cell blank
  - Missing keys are treated as None (blank cell) — you do not need to
    supply an entry for every possible combination

Typical loop to populate:
    data = {}
    for period, cols in periods.items():
        for run_type, run in cols:
            for string_num, detectors in strings.items():
                for ged_name, _ in detectors:
                    for usability_type in ("E", "P"):
                        value = blah
                        data[(string_num, ged_name, period, run,
                              run_type, usability_type)] = value

================================================================================
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Visual stuff — edit here to restyle the sheet if you want
# ---------------------------------------------------------------------------

# One hex color per period label; add entries for new periods as needed.
PERIOD_COLOURS: dict[str, str] = {
    "p16": "4472C4",  # blue
    "p18": "70AD47",  # green
    "p19": "ED7D31",  # orange
    "p20": "C44444",  # red
    "p21": "7100B7",  # purple
    "p22": "BB00BB",  # pink
}
DEFAULT_PERIOD_COLOUR = "808080"  # grey fallback for unlisted periods

CAL_HEADER_ROW_FILL = "98FAE9"  # teal — cal run header
PHY_HEADER_ROW_FILL = "FF42A1"  # pink — phy run header
STRING_FILL_ODD = "F2F2F2"  # light grey — alternating string groups
STRING_FILL_EVEN = "E0E0E0"  # mid grey
ESCALE_FILL = "F6D7B8"  # Escale row marker
PSD_FILL = "95B3DE"  # PSD row marker
WHITE = "FFFFFF"

# Conditional formatting colours
CF_GREEN = "CEEED0"  # on
CF_AMBER = "FCEBA5"  # ac
CF_RED = "F7C9CF"  # off
CF_LIGHT_BLUE = "B0CBFFFF"  # valid
CF_BLUE = "8A9FC5FF"  # present
CF_DARK_BLUE = "667590FF"  # missing

# ---------------------------------------------------------------------------
# Excel sheet functions
# ---------------------------------------------------------------------------


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border(left=None, right=None, top=None, bottom=None) -> Border:
    def s(style):
        return Side(style=style) if style else Side()

    return Border(left=s(left), right=s(right), top=s(top), bottom=s(bottom))


def add_summary_rows(
    work_sheet,
    strings: dict,
    periods: dict,
    col_index: dict,
    livetimes: dict,
) -> None:
    """
    Append livetime and exposure summary rows below the detector data.

      Livetime [days]              — actual values, phy columns only
      Exposure ON    [kg·yr]       — Excel SUMPRODUCT formula
      Exposure AC    [kg·yr]       — Excel SUMPRODUCT formula
      Exposure OFF   [kg·yr]       — Excel SUMPRODUCT formula
      Exposure PSD valid+ON [kg·yr] — detectors with E=on AND P=valid
    """
    total_dets = sum(len(dets) for dets in strings.values())
    last_row = 2 + total_dets * 2  # last P row of the data block
    last_e_row = last_row - 1  # last E row
    first_data = 3

    summary_start = last_row + 3
    livetime_row = summary_start
    on_row = summary_start + 1
    ac_row = summary_start + 2
    off_row = summary_start + 3
    psd_row = summary_start + 4

    # colours
    value_fills = {
        livetime_row: "EBF3FB",
        on_row: "E2EFDA",  # green tint
        ac_row: "FFF2CC",  # amber tint
        off_row: "FCE4D6",  # red tint
        psd_row: "E8DAEF",  # purple tint
    }
    labels = {
        livetime_row: "Livetime [days]",
        on_row: "Exposure ON [kg·yr]",
        ac_row: "Exposure AC [kg·yr]",
        off_row: "Exposure OFF [kg·yr]",
        psd_row: "Exposure PSD valid+ON [kg·yr]",
    }
    label_fill = _fill("D9E1F2")  # pale blue

    for row, label in labels.items():
        cell = work_sheet.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, size=10)
        cell.fill = _fill(label_fill.fgColor.rgb)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _border(left="medium", right="thin", top="thin", bottom="thin")
        work_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    period_list = [period for period in periods]

    for period in period_list:
        for run_type, run in periods[period]:
            if run_type != "phy":
                continue
            livetime_s = livetimes.get((period, run))
            if livetime_s is None:
                continue

            col_idx = col_index.get((period, run_type, run))
            if col_idx is None:
                continue
            col_letter = get_column_letter(col_idx)
            livetime_days = livetime_s / 60 / 60 / 24

            # Livetime cell — actual value, overwritable in Excel
            cell = work_sheet.cell(
                row=livetime_row, column=col_idx, value=round(livetime_days, 2)
            )
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"
            cell.fill = _fill(value_fills[livetime_row])
            cell.border = _border(right="thin", top="thin", bottom="thin")

            # Reference to the livetime cell in the same column (days → years)
            livetime_ref = f"{col_letter}{livetime_row}"

            # Full data range (E + P rows) used for on/ac/off formulas
            full = f"C{first_data}:C{last_row}"
            x_full = f"{col_letter}{first_data}:{col_letter}{last_row}"

            # Separate E-only and P-only ranges for the PSD formula (same length)
            x_e = f"{col_letter}{first_data}:{col_letter}{last_e_row}"
            x_p = f"{col_letter}{first_data+1}:{col_letter}{last_row}"
            c_e = f"C{first_data}:C{last_e_row}"

            # Excel formulas
            def _exposure(
                usability: str,
                full=full,
                first_data=first_data,
                x_full=x_full,
                livetime_ref=livetime_ref,
            ) -> str:
                # Sum masses (g->kg) of E rows where usability matches, x livetime (yr)
                return (
                    f"=SUMPRODUCT("
                    f"(MOD(ROW({full})-{first_data},2)=0)*"
                    f'({x_full}="{usability}")*'
                    f"{full}"
                    f")/1000*{livetime_ref}/365.25"
                )

            def _psd_exposure(
                c_e=c_e,
                first_data=first_data,
                x_e=x_e,
                x_p=x_p,
                livetime_ref=livetime_ref,
            ) -> str:
                # E row = "on" AND corresponding P row = "valid"
                return (
                    f"=SUMPRODUCT("
                    f"(MOD(ROW({c_e})-{first_data},2)=0)*"
                    f'({x_e}="on")*'
                    f'({x_p}="valid")*'
                    f"{c_e}"
                    f")/1000*{livetime_ref}/365.25"
                )

            for row, formula in [
                (on_row, _exposure("on")),
                (ac_row, _exposure("ac")),
                (off_row, _exposure("off")),
                (psd_row, _psd_exposure()),
            ]:
                cell = work_sheet.cell(row=row, column=col_idx, value=formula)
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0.000"
                cell.fill = _fill(value_fills[row])
                cell.border = _border(right="thin", top="thin", bottom="thin")


def make_excel(
    strings: dict,
    periods: dict,
    data: dict,
    output_path: str = "dashboard_output.xlsx",
    livetimes: dict = None,
) -> None:
    """
    Generate the usability dashboard Excel file.

    Parameters
    ----------
    strings     : see module docstring — detector layout per string
    periods     : see module docstring — (run_type, run) columns per period
    data        : see module docstring — usability values
    output_path : path to write the .xlsx file
    livetimes   : optional {(period, run): livetime_in_seconds} — if supplied,
                  summary exposure rows are appended below the detector data
    """
    work_book = openpyxl.Workbook()
    work_sheet = work_book.active
    work_sheet.title = "Usability"

    period_list = list(periods.keys())

    # Build (period, run_type, run) -> 1-based column index
    col_index: dict[tuple, int] = {}
    col = 5  # A=String, B=GED, C=Mass, D=E/P type
    for period in period_list:
        for run_type, run in periods[period]:
            col_index[(period, run_type, run)] = col
            col += 1

    # row 1 period headers
    work_sheet.row_dimensions[1].height = 17

    for period in period_list:
        cols = [col_index[(period, runtype, run)] for runtype, run in periods[period]]
        start_col, end_col = cols[0], cols[-1]
        color = PERIOD_COLOURS.get(period, DEFAULT_PERIOD_COLOUR)
        cell = work_sheet.cell(row=1, column=start_col, value=period)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border(
            left="medium", right="medium", top="medium", bottom="medium"
        )
        if end_col > start_col:
            work_sheet.merge_cells(
                start_row=1, start_column=start_col, end_row=1, end_column=end_col
            )

    # row 2 column headers
    work_sheet.row_dimensions[2].height = 50

    for label, col in [("String", 1), ("GED", 2), ("Mass [g]", 3)]:
        cell = work_sheet.cell(row=2, column=col, value=label)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 1:
            cell.border = _border(left="medium")

    for period in period_list:
        done_boarder = False
        for run_type, run in periods[period]:
            col = col_index[(period, run_type, run)]
            cell = work_sheet.cell(row=2, column=col, value=f"{run_type} {run}")
            if not done_boarder:
                cell.border = _border(left="medium")
                done_boarder = True
            cell.font = Font(bold=True, size=11)
            if run_type == "cal":
                cell.fill = _fill(CAL_HEADER_ROW_FILL)
            else:
                cell.fill = _fill(PHY_HEADER_ROW_FILL)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", text_rotation=90
            )

    # column widths
    work_sheet.column_dimensions["A"].width = 8
    work_sheet.column_dimensions["B"].width = 10
    work_sheet.column_dimensions["C"].width = 8
    work_sheet.column_dimensions["D"].width = 3
    for col in col_index.values():
        work_sheet.column_dimensions[get_column_letter(col)].width = 5

    # data rows
    current_row = 3
    for str_idx, string_num in enumerate(sorted(strings.keys())):
        detectors = strings[string_num]
        str_fill = _fill(STRING_FILL_ODD if str_idx % 2 == 0 else STRING_FILL_EVEN)
        string_start_row = current_row
        string_end_row = current_row + len(detectors) * 2 - 1

        for det_idx, (ged_name, mass) in enumerate(detectors):
            e_row = current_row
            p_row = current_row + 1
            first = det_idx == 0
            last = det_idx == len(detectors) - 1

            # Column A — string number
            a = work_sheet.cell(row=e_row, column=1, value=string_num)
            a.fill = str_fill
            a.font = Font(size=11, color="000000")
            a.alignment = Alignment(horizontal="center", vertical="center")
            a.border = _border(
                left="medium",
                right="thin",
                top="medium" if first else "thin",
                bottom="medium" if last else "thin",
            )

            # Column B — GED name (merged E+P rows)
            b = work_sheet.cell(row=e_row, column=2, value=ged_name)
            b.font = Font(size=11)
            b.alignment = Alignment(horizontal="center", vertical="center")
            b.border = _border(
                top="medium" if first else "thin", bottom="medium" if last else "thin"
            )
            work_sheet.merge_cells(
                start_row=e_row, start_column=2, end_row=p_row, end_column=2
            )

            # Column C — mass (merged E+P rows)
            c = work_sheet.cell(row=e_row, column=3, value=mass)
            c.font = Font(size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border(
                top="medium" if first else "thin", bottom="medium" if last else "thin"
            )
            work_sheet.merge_cells(
                start_row=e_row, start_column=3, end_row=p_row, end_column=3
            )

            # Column D — usability type label (E / P)
            for row_offset, usability_type in [(0, "E"), (1, "P")]:
                row = e_row + row_offset
                d = work_sheet.cell(row=row, column=4, value=usability_type)
                d.fill = _fill(ESCALE_FILL if usability_type == "E" else PSD_FILL)
                d.font = Font(size=11)
                d.alignment = Alignment(horizontal="center", vertical="center")
                top = "medium" if (first and usability_type == "E") else "thin"
                bot = "medium" if (last and usability_type == "P") else "thin"
                d.border = _border(left="thin", top=top, bottom=bot)

            # Data columns
            for period in period_list:
                period_cols = periods[period]
                for col_pos, (run_type, run) in enumerate(period_cols):
                    col_idx = col_index[(period, run_type, run)]
                    is_first_col = col_pos == 0

                    reason = data.get(
                        (string_num, ged_name, period, run, run_type, "reason")
                    )

                    for row_offset, usability_type in [(0, "E"), (1, "P")]:
                        row = e_row + row_offset
                        cell = work_sheet.cell(row=row, column=col_idx)
                        value = data.get(
                            (
                                string_num,
                                ged_name,
                                period,
                                run,
                                run_type,
                                usability_type,
                            )
                        )
                        cell.value = value
                        cell.font = Font(size=9, color="000000")
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        top = "medium" if (first and usability_type == "E") else "thin"
                        bot = "medium" if (last and usability_type == "P") else "thin"
                        cell.border = _border(
                            left="medium" if is_first_col else "thin",
                            right="thin",
                            top=top,
                            bottom=bot,
                        )

                        # Comments: reason on E row, PSD cut statuses on P row
                        if usability_type == "E" and reason:
                            cell.comment = Comment(
                                reason, author="legend-datasets", height=60, width=200
                            )
                        if usability_type == "P":
                            psd_note = data.get(
                                (
                                    string_num,
                                    ged_name,
                                    period,
                                    run,
                                    run_type,
                                    "PSD_note",
                                )
                            )
                            if psd_note:
                                n_lines = psd_note.count("\n") + 1
                                cell.comment = Comment(
                                    psd_note,
                                    author="legend-datasets",
                                    height=15 * n_lines + 10,
                                    width=160,
                                )

            current_row += 2

        # Merge string number vertically over all its detector rows
        if string_end_row > string_start_row:
            work_sheet.merge_cells(
                start_row=string_start_row,
                start_column=1,
                end_row=string_end_row,
                end_column=1,
            )

    # conditional formatting on the data range
    total_dets = sum(len(dets) for dets in strings.values())
    last_data_row = 2 + total_dets * 2
    first_data_col = get_column_letter(5)  # always column E
    last_data_col = get_column_letter(max(col_index.values()))
    cf_range = f"{first_data_col}3:{last_data_col}{last_data_row}"
    ref = f"{first_data_col}3"  # top-left cell — Excel adjusts relatively

    for values, hex_color in [
        (["on", "valid"], CF_GREEN),
        (["ac"], CF_AMBER),
        (["off"], CF_RED),
        (["present"], CF_BLUE),
        (["missing"], CF_DARK_BLUE),
    ]:
        fill = PatternFill(
            start_color=hex_color, end_color=hex_color, fill_type="solid"
        )
        condition = ",".join(f'{ref}="{v}"' for v in values)
        work_sheet.conditional_formatting.add(
            cf_range,
            FormulaRule(formula=[f"OR({condition})"], fill=fill),
        )

    # Freeze rows 1-2 and columns A-D
    work_sheet.freeze_panes = "E3"

    if livetimes:
        add_summary_rows(work_sheet, strings, periods, col_index, livetimes)

    work_book.save(output_path)


# ---------------------------------------------------------------------------
# QCP check constants
# ---------------------------------------------------------------------------

QCP_TRUE_FILL = "CEEED0"  # green — all checks passed
QCP_FALSE_FILL = "F7C9CF"  # red   — at least one non-PSD check failed
QCP_PSD_FILL = "9DC3E6"  # blue  — only PSD failed
QCP_NULL_FILL = "E8E8E8"  # grey  — no data


QCP_CAL_CHECKS = [
    ("FEP_gain_stab", "FEP stab", "F6D7B8"),
    ("fwhm_ok", "FWHM", "F6D7B8"),
    ("npeak", "Npeak", "F6D7B8"),
    ("const_stab", "Const", "F6D7B8"),
    ("PSD", "PSD", "95B3DE"),
]
QCP_PHY_CHECKS = [
    ("baseln_spike", "BL spike", "FFD0E8"),
    ("baseln_stab", "BL stab", "FFD0E8"),
    ("pulser_stab", "Pulser", "FFD0E8"),
]


def _qcp_result(det_qcp: dict, run_type: str) -> tuple[str | None, list[str]]:
    """
    Check all checks for a ged for a run type.

    Returns (result, failed_checks) where result is "pass", "fail", or None.
    None means all checks were null (no data available).
    """
    checks = QCP_CAL_CHECKS if run_type == "cal" else QCP_PHY_CHECKS
    section = det_qcp.get(run_type, {})
    checks_non_null = {
        k[0]: section[k[0]]
        for k in checks
        if k[0] in section and section[k[0]] is not None
    }
    if not checks_non_null:
        return None, []
    failed = [k for k, v in checks_non_null.items() if v is False]
    return ("fail" if failed else "pass"), failed


def make_qcp_sheet(
    work_book_path: str,
    strings: dict,
    periods: dict,
    qcp_data: dict,
) -> None:
    """
    Add a 'QCP Summary' sheet to an existing workbook at wb_path.

    One row per detector.  Columns:
    cal r000, phy r000, cal r001, phy r001, ...
    Each cell shows "pass" or "fail" if
    any fail — with a comment listing the failing check names.

    Parameters
    ----------
    wb_path  : path to an existing .xlsx file (produced by make_excel)
    strings  : same dict passed to make_excel
    periods  : same dict passed to make_excel
    qcp_data : {period: {run: {detector: {"cal": {...}, "phy": {...}}}}}
               as returned by read_qcp.get_qcp_data()
    """
    work_book = load_workbook(work_book_path)
    work_sheet = work_book.create_sheet("QCP Summary")

    col_index: dict[tuple, int] = {}
    col = 4
    for period in periods:
        for run_type, run in periods[period]:
            col_index[(period, run_type, run)] = col
            col += 1

    # row 1: period headers
    work_sheet.row_dimensions[1].height = 17

    for period in periods:
        cols = [col_index[(period, rt, r)] for rt, r in periods[period]]
        start_col, end_col = cols[0], cols[-1]
        color = PERIOD_COLOURS.get(period, DEFAULT_PERIOD_COLOUR)
        cell = work_sheet.cell(row=1, column=start_col, value=period)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border(
            left="medium", right="medium", top="medium", bottom="medium"
        )
        if end_col > start_col:
            work_sheet.merge_cells(
                start_row=1, start_column=start_col, end_row=1, end_column=end_col
            )

    # row 2: column headers
    work_sheet.row_dimensions[2].height = 50

    for label, col in [("String", 1), ("GED", 2), ("Mass [g]", 3)]:
        cell = work_sheet.cell(row=2, column=col, value=label)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 1:
            cell.border = _border(left="medium")

    for period in periods:
        done_border = False
        for run_type, run in periods[period]:
            col = col_index[(period, run_type, run)]
            cell = work_sheet.cell(row=2, column=col, value=f"{run_type} {run}")
            cell.font = Font(bold=True, size=11)
            cell.fill = _fill(
                CAL_HEADER_ROW_FILL if run_type == "cal" else PHY_HEADER_ROW_FILL
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", text_rotation=90
            )
            if not done_border:
                cell.border = _border(left="medium")
                done_border = True

    # column widths
    work_sheet.column_dimensions["A"].width = 8
    work_sheet.column_dimensions["B"].width = 10
    work_sheet.column_dimensions["C"].width = 8
    for col in col_index.values():
        work_sheet.column_dimensions[get_column_letter(col)].width = 5

    # data rows
    current_row = 3
    for str_idx, string_num in enumerate(sorted(strings.keys())):
        detectors = strings[string_num]
        str_fill = _fill(STRING_FILL_ODD if str_idx % 2 == 0 else STRING_FILL_EVEN)
        string_start_row = current_row
        string_end_row = current_row + len(detectors) - 1

        for det_idx, (ged_name, mass) in enumerate(detectors):
            row = current_row
            first = det_idx == 0
            last = det_idx == len(detectors) - 1

            # Column A — string number (written every row; merged at end)
            a = work_sheet.cell(row=row, column=1, value=string_num)
            a.fill = str_fill
            a.font = Font(size=11)
            a.alignment = Alignment(horizontal="center", vertical="center")
            a.border = _border(
                left="medium",
                right="thin",
                top="medium" if first else "thin",
                bottom="medium" if last else "thin",
            )

            # Column B — GED name
            b = work_sheet.cell(row=row, column=2, value=ged_name)
            b.font = Font(size=11)
            b.alignment = Alignment(horizontal="center", vertical="center")
            b.border = _border(
                top="medium" if first else "thin", bottom="medium" if last else "thin"
            )

            # Column C — mass
            c_cell = work_sheet.cell(row=row, column=3, value=mass)
            c_cell.font = Font(size=11)
            c_cell.alignment = Alignment(horizontal="center", vertical="center")
            c_cell.border = _border(
                top="medium" if first else "thin", bottom="medium" if last else "thin"
            )

            # Data columns
            for period in periods:
                done_border = False
                for run_type, run in periods[period]:
                    c_idx = col_index[(period, run_type, run)]
                    det_qcp = qcp_data.get(period, {}).get(run, {}).get(ged_name, {})
                    result, failed = _qcp_result(det_qcp, run_type)

                    if result == "pass":
                        fill_hex, display = QCP_TRUE_FILL, "pass"
                    elif result == "fail" and failed == ["PSD"]:
                        fill_hex, display = QCP_PSD_FILL, "fail"
                    elif result == "fail":
                        fill_hex, display = QCP_FALSE_FILL, "fail"
                    else:
                        fill_hex, display = QCP_NULL_FILL, None

                    cell = work_sheet.cell(row=row, column=c_idx, value=display)
                    cell.fill = _fill(fill_hex)
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = _border(
                        left="medium" if not done_border else "thin",
                        right="thin",
                        top="medium" if first else "thin",
                        bottom="medium" if last else "thin",
                    )
                    done_border = True

                    if failed:
                        note = "Failed:\n" + "\n".join(f"  {k}" for k in failed)
                        cell.comment = Comment(
                            note,
                            author="qcp",
                            height=15 * (len(failed) + 1) + 10,
                            width=160,
                        )

            current_row += 1

        # Merge string column vertically over all detectors in the string
        if string_end_row > string_start_row:
            work_sheet.merge_cells(
                start_row=string_start_row,
                start_column=1,
                end_row=string_end_row,
                end_column=1,
            )

    work_sheet.freeze_panes = "D3"
    work_book.save(work_book_path)


def _make_qcp_detail_sheet(
    work_book,
    sheet_name: str,
    run_type_filter: str,
    checks: list,
    strings: dict,
    periods: dict,
    qcp_data: dict,
) -> None:
    work_sheet = work_book.create_sheet(sheet_name)
    n_checks = len(checks)

    # Only include columns whose run_type matches phy or cal
    col_index: dict[tuple, int] = {}
    col = 5
    for period in periods:
        for run_type, run in periods[period]:
            if run_type == run_type_filter:
                col_index[(period, run)] = col
                col += 1

    # row 1: period headers
    work_sheet.row_dimensions[1].height = 17

    for period in periods:
        period_cols = [
            col_index[(period, run)]
            for run_type, run in periods[period]
            if run_type == run_type_filter and (period, run) in col_index
        ]
        if not period_cols:
            continue
        start_col, end_col = period_cols[0], period_cols[-1]
        color = PERIOD_COLOURS.get(period, DEFAULT_PERIOD_COLOUR)
        cell = work_sheet.cell(row=1, column=start_col, value=period)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border(
            left="medium", right="medium", top="medium", bottom="medium"
        )
        if end_col > start_col:
            work_sheet.merge_cells(
                start_row=1, start_column=start_col, end_row=1, end_column=end_col
            )

    # row 2: column headers
    work_sheet.row_dimensions[2].height = 50

    for label, col in [("String", 1), ("GED", 2), ("Mass [g]", 3), ("Check", 4)]:
        cell = work_sheet.cell(row=2, column=col, value=label)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 1:
            cell.border = _border(left="medium")

    header_fill = (
        CAL_HEADER_ROW_FILL if run_type_filter == "cal" else PHY_HEADER_ROW_FILL
    )
    for period in periods:
        first_in_period = True
        for run_type, run in periods[period]:
            if run_type != run_type_filter:
                continue
            col = col_index[(period, run)]
            cell = work_sheet.cell(row=2, column=col, value=run)
            cell.font = Font(bold=True, size=11)
            cell.fill = _fill(header_fill)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", text_rotation=90
            )
            if first_in_period:
                cell.border = _border(left="medium")
                first_in_period = False

    # column widths
    work_sheet.column_dimensions["A"].width = 8
    work_sheet.column_dimensions["B"].width = 10
    work_sheet.column_dimensions["C"].width = 8
    work_sheet.column_dimensions["D"].width = 8
    for col in col_index.values():
        work_sheet.column_dimensions[get_column_letter(col)].width = 5

    # data rows
    current_row = 3
    for str_idx, string_num in enumerate(sorted(strings.keys())):
        detectors = strings[string_num]
        str_fill = _fill(STRING_FILL_ODD if str_idx % 2 == 0 else STRING_FILL_EVEN)
        string_start_row = current_row
        string_end_row = current_row + len(detectors) * n_checks - 1

        for det_idx, (ged_name, mass) in enumerate(detectors):
            det_start = current_row
            det_end = current_row + n_checks - 1
            first_det = det_idx == 0
            last_det = det_idx == len(detectors) - 1

            # Column A — string number (written per detector; string-level merge at end)
            a = work_sheet.cell(row=det_start, column=1, value=string_num)
            a.fill = str_fill
            a.font = Font(size=11)
            a.alignment = Alignment(horizontal="center", vertical="center")
            a.border = _border(
                left="medium",
                right="thin",
                top="medium" if first_det else "thin",
                bottom="medium" if last_det else "thin",
            )

            # Column B — GED (merged over check rows)
            b = work_sheet.cell(row=det_start, column=2, value=ged_name)
            b.font = Font(size=11)
            b.alignment = Alignment(horizontal="center", vertical="center")
            b.border = _border(
                top="medium" if first_det else "thin",
                bottom="medium" if last_det else "thin",
            )
            work_sheet.merge_cells(
                start_row=det_start, start_column=2, end_row=det_end, end_column=2
            )

            # Column C — mass (merged over check rows)
            c_cell = work_sheet.cell(row=det_start, column=3, value=mass)
            c_cell.font = Font(size=11)
            c_cell.alignment = Alignment(horizontal="center", vertical="center")
            c_cell.border = _border(
                top="medium" if first_det else "thin",
                bottom="medium" if last_det else "thin",
            )
            work_sheet.merge_cells(
                start_row=det_start, start_column=3, end_row=det_end, end_column=3
            )

            # One row per check
            for chk_idx, (yaml_key, label, check_fill_hex) in enumerate(checks):
                row = det_start + chk_idx
                first_chk = chk_idx == 0
                last_chk = chk_idx == n_checks - 1
                top = "medium" if (first_det and first_chk) else "thin"
                bot = "medium" if (last_det and last_chk) else "thin"

                # Column D — check label
                d = work_sheet.cell(row=row, column=4, value=label)
                d.fill = _fill(check_fill_hex)
                d.font = Font(size=9)
                d.alignment = Alignment(horizontal="center", vertical="center")
                d.border = _border(left="thin", top=top, bottom=bot)

                # Data columns
                for period in periods:
                    first_in_period = True
                    for run_type, run in periods[period]:
                        if run_type != run_type_filter:
                            continue
                        c_idx = col_index[(period, run)]
                        det_qcp = (
                            qcp_data.get(period, {}).get(run, {}).get(ged_name, {})
                        )
                        value = det_qcp.get(run_type_filter, {}).get(yaml_key)

                        if value is True:
                            fill_hex, display = QCP_TRUE_FILL, "pass"
                        elif value is False:
                            fill_hex, display = QCP_FALSE_FILL, "fail"
                        else:
                            fill_hex, display = QCP_NULL_FILL, None

                        cell = work_sheet.cell(row=row, column=c_idx, value=display)
                        cell.fill = _fill(fill_hex)
                        cell.font = Font(size=8)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        cell.border = _border(
                            left="medium" if first_in_period else "thin",
                            right="thin",
                            top=top,
                            bottom=bot,
                        )
                        first_in_period = False

            current_row += n_checks

        if string_end_row > string_start_row:
            work_sheet.merge_cells(
                start_row=string_start_row,
                start_column=1,
                end_row=string_end_row,
                end_column=1,
            )

    work_sheet.freeze_panes = "E3"


def make_qcp_sheets_detailed(
    wb_path: str,
    strings: dict,
    periods: dict,
    qcp_data: dict,
) -> None:
    """
    Add 'QCP Cal' and 'QCP Phy' sheets to an existing workbook.

    Each sheet shows one row per check per detector with pass/fail/null
    coloring.
    """
    wb = load_workbook(wb_path)
    _make_qcp_detail_sheet(
        wb, "QCP Cal", "cal", QCP_CAL_CHECKS, strings, periods, qcp_data
    )
    _make_qcp_detail_sheet(
        wb, "QCP Phy", "phy", QCP_PHY_CHECKS, strings, periods, qcp_data
    )
    wb.save(wb_path)
