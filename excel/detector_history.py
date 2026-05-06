"""
Build per-detector transition histories from either the on-disk validity + config files, or from the Excel dashboard.

A Transition represents a genuine change in a detector's usability at a
specific (period, run, run_type).  The ordered list of Transitions for a
detector fully describes its history over the tracked periods.
"""

from __future__ import annotations

from pathlib import Path

import yaml
from dbetto import TextDB
from openpyxl import load_workbook

# Type alias: one list of Transitions per detector name.
History = dict[str, list["Transition"]]


class Transition:
    """A class to hold a bunch of info for a change in a detector's usability at a specific run."""

    ged: str  #
    period: str  # pXX
    run: str  # rXXX
    run_type: str  # "cal" or "phy"
    timestamp: str  # runinfo start_key
    usability: str  # on/ac/off
    reason: str | None = None  # string with the stated reason
    source_file: str | None = None  # config basename; None for Excel transitions
    # or baseline (pre-window) disk values


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _ordered_entries(periods: dict, runinfo: dict) -> list[tuple[str, str, str]]:
    """Return all (period, run, run_type) tuples in the order they appear in `periods`, filtered to entries that exist in runinfo."""
    result = []
    for period, cols in periods.items():
        period_ri = runinfo.get(period, {})
        for run_type, run in cols:
            if run in period_ri and run_type in period_ri[run]:
                result.append((period, run, run_type))
    return result


def _all_runinfo_entries(runinfo: dict) -> list[tuple[str, str, str]]:
    """Return all (period, run, run_type) tuples across every period in runinfo, sorted chronologically by start_key.

    Used by build_from_disk so that prev_val tracking is always correct
    regardless of which periods the Excel sheet covers.
    """
    entries = []
    for period, runs in runinfo.items():
        for run, run_data in runs.items():
            if not isinstance(run_data, dict):
                continue
            for run_type in ("cal", "phy"):
                if run_type in run_data and "start_key" in run_data[run_type]:
                    ts = run_data[run_type]["start_key"]
                    entries.append((period, run, run_type, ts))
    entries.sort(key=lambda x: x[3])
    return [(period, run, runtype) for period, run, runtype, _ in entries]


def _find_source_and_reason(
    statuses_dir: Path,
    validity: list,
    timestamp: str,
    ged: str,
) -> tuple[str | None, str | None]:
    """Scan validity entries at `timestamp` to find the config file that explicitly sets `ged`.

    Returns (source_file_basename, reason).  When multiple config files at the
    same timestamp contain `ged`, the last one wins (mirrors TextDB append order).
    """
    result_file = None
    result_reason = None

    for entry in validity:
        if entry["valid_from"] != timestamp:
            continue
        for config_name in entry.get("apply", []):
            p = statuses_dir / config_name
            if not p.exists():
                continue
            with open(p) as f:
                cfg = yaml.safe_load(f) or {}
            if ged in cfg:
                result_file = config_name
                result_reason = cfg[ged].get("reason")

    return result_file, result_reason


# ---------------------------------------------------------------------------
# Build from disk
# ---------------------------------------------------------------------------


def build_from_disk(
    datasets: Path,
    strings: dict,
    excel_periods: dict,
    runinfo: dict,
) -> tuple[dict, History]:
    """
    Build a per-detector transition list from the on-disk validity + config files.

    Walks ALL periods in runinfo chronologically so that prev_val tracking is
    always correct — regardless of which periods the Excel sheet covers.

    The returned usability dict is filtered to `excel_periods` only, so it can
    be compared directly with the usability dict from build_from_excel.

    source_file and reason are read from the config file that explicitly sets
    the value at the transition timestamp.  A Transition with source_file=None
    means the value is inherited — no config file at this exact timestamp.
    """
    statuses_dir = datasets / "statuses"
    statuses = TextDB(str(statuses_dir))

    with open(statuses_dir / "validity.yaml") as f:
        validity = yaml.safe_load(f) or []

    all_geds = [ged for snum in sorted(strings) for ged, _ in strings[snum]]
    all_entries = _all_runinfo_entries(runinfo)
    excel_events_set = set(_ordered_entries(excel_periods, runinfo))

    result: History = {ged: [] for ged in all_geds}
    prev_usab: dict[str, str | None] = {ged: None for ged in all_geds}
    usability: dict = {ged: {} for ged in all_geds}

    for period, run, run_type in all_entries:
        start_key = runinfo[period][run][run_type]["start_key"]
        status = statuses.on(start_key, system=run_type)

        for ged in all_geds:
            usab = status.get(ged, {}).get("usability")

            # Populate usability dict only for Excel-period events
            if (period, run, run_type) in excel_events_set:
                usability[ged].setdefault(period, {}).setdefault(run, {})[
                    run_type
                ] = usab

            if usab is not None and usab != prev_usab[ged]:
                source, reason = _find_source_and_reason(
                    statuses_dir, validity, start_key, ged
                )
                result[ged].append(
                    Transition(
                        ged=ged,
                        period=period,
                        run=run,
                        run_type=run_type,
                        timestamp=start_key,
                        usability=usab,
                        reason=reason,
                        source_file=source,
                    )
                )

            prev_usab[ged] = usab

    return usability, result


# ---------------------------------------------------------------------------
# Build from Excel
# ---------------------------------------------------------------------------


def build_from_excel(
    xlsx_path: str,
    strings: dict,
    periods: dict,
    runinfo: dict,
    prev_usab_seed: dict[str, str | None] | None = None,
) -> History:
    """Build a per-detector transition list from the Excel usability matrix.

    Walks events in period-column order, reading each cell value.  Records a
    Transition wherever the value changes relative to the previous event.
    Cell comments are read as the reason for that transition.

    prev_usab_seed, if provided, is used to initialise the per-detector
    "last seen" value before the first Excel event. Pass the on-disk
    usability at the first Excel entry
    """
    work_book = load_workbook(xlsx_path, data_only=True)
    work_sheet = work_book["Usability"]

    # (period, run_type, run) -> column number
    col_index: dict[tuple, int] = {}
    col = 5  # start from column 5 (string, ged, mass, P/E, X)
    for period in periods:
        for run_type, run in periods[period]:
            col_index[(period, run_type, run)] = col
            col += 1

    # (string_num, ged) -> E-row number  (two rows per detector: Escale then PSD)
    row_index: dict[tuple, int] = {}
    row = 3  # start from row 3 (pXX, cal/phy rYYY, X)
    for string_num in sorted(strings.keys()):
        for ged, _ in strings[string_num]:
            row_index[(string_num, ged)] = row
            row += 2

    all_geds_ordered = [
        (string_num, ged)
        for string_num in sorted(strings)
        for ged, _ in strings[string_num]
    ]
    entries = _ordered_entries(periods, runinfo)

    result: History = {ged: [] for _, ged in all_geds_ordered}
    prev_usab: dict[str, str | None] = dict(prev_usab_seed) if prev_usab_seed else {}
    usability = {}

    for string_num, ged in all_geds_ordered:
        if ged not in usability.keys():
            usability[ged] = {}
        for period, run, run_type in entries:
            start_key = runinfo[period][run][run_type]["start_key"]

            e_row = row_index.get((string_num, ged))
            if e_row is None:
                continue

            cell = work_sheet.cell(row=e_row, column=col_index[(period, run_type, run)])
            usab = cell.value

            if period not in usability[ged].keys():
                usability[ged][period] = {}
            if run not in usability[ged][period].keys():
                usability[ged][period][run] = {}
            usability[ged][period][run][run_type] = usab

            reason = None
            if cell.comment:
                text = cell.comment.text.strip()
                if text:
                    reason = text

            if usab != prev_usab.get(ged):
                result[ged].append(
                    Transition(
                        ged=ged,
                        period=period,
                        run=run,
                        run_type=run_type,
                        timestamp=start_key,
                        usability=usab,
                        reason=reason,
                        source_file=None,
                    )
                )

            prev_usab[ged] = usab

    return usability, result
